import numpy as np
from skimage import io
from tkinter import Tk, filedialog, Label, Button, Scale, HORIZONTAL, messagebox, ttk, Frame
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
import os

# === Функции работы с матрицами и Excel ===

def save_matrix_to_excel(matrices, filename):
    """Сохраняет словарь {имя: матрица} в Excel файл"""
    wb = Workbook()
    for name, matrix in matrices.items():
        ws = wb.create_sheet(title=name[:31])  # Excel ограничение — 31 символ
        m, n = matrix.shape
        for i in range(m):
            row_values = matrix[i, :].tolist()
            ws.append(row_values)
    # Удаляем дефолтный лист
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    wb.save(filename)


def compress_image_svd(img, k, save_dir, basename):
    """Компрессия изображения через SVD + сохранение матриц в Excel"""
    # Проверяем ч/б или RGB
    if img.ndim == 2:
        num_channels = 1
    elif img.ndim == 3 and img.shape[2] >= 3:
        if np.allclose(img[:, :, 0], img[:, :, 1]) and np.allclose(img[:, :, 1], img[:, :, 2]):
            img = img[:, :, 0]
            num_channels = 1
        else:
            num_channels = 3
    else:
        num_channels = 1

    new_img = []
    num_loops = 3 if num_channels > 1 else 1

    for i in range(num_loops):
        channel = img if num_channels == 1 else img[:, :, i]
        channel = channel.astype(float)

        U, S, Vt = np.linalg.svd(channel, full_matrices=False)
        U_k, S_k, Vt_k = U[:, :k], np.diag(S[:k]), Vt[:k, :]

        matrices = {
            f"U_full_ch{i+1}": U,
            f"S_full_ch{i+1}": np.diag(S),
            f"Vt_full_ch{i+1}": Vt,
            f"U_k={k}_ch{i+1}": U_k,
            f"S_k={k}_ch{i+1}": S_k,
            f"Vt_k={k}_ch{i+1}": Vt_k
        }

        excel_path = os.path.join(save_dir, f"{basename}_channel{i+1}_matrices.xlsx")
        save_matrix_to_excel(matrices, excel_path)

        compressed_channel = U_k @ S_k @ Vt_k
        new_img.append(compressed_channel)

    img_compressed = np.stack(new_img, axis=2) if num_channels > 1 else new_img[0]
    img_compressed = np.clip(img_compressed, 0, 255).astype(np.uint8)
    return img_compressed


def compress_excel_svd(filepath, k, save_dir, basename):
    """Компрессия Excel таблицы через SVD"""
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    # Читаем таблицу в numpy
    data = []
    for row in ws.iter_rows(values_only=True):
        row_values = []
        for v in row:
            try:
                row_values.append(float(v))
            except (TypeError, ValueError):
                row_values.append(0.0)
        data.append(row_values)

    matrix = np.array(data, dtype=float)

    # SVD
    U, S, Vt = np.linalg.svd(matrix, full_matrices=False)
    U_k, S_k, Vt_k = U[:, :k], np.diag(S[:k]), Vt[:k, :]

    # Восстановленная таблица
    compressed_matrix = U_k @ S_k @ Vt_k

    # Сохраняем в Excel
    result_path = os.path.join(save_dir, f"{basename}_compressed_k={k}.xlsx")
    wb_out = Workbook()
    ws_out = wb_out.active
    for row in compressed_matrix:
        ws_out.append(list(row))
    wb_out.save(result_path)

    # Матрицы отдельно
    matrices = {
        "U_full": U,
        "S_full": np.diag(S),
        "Vt_full": Vt,
        f"U_k={k}": U_k,
        f"S_k={k}": S_k,
        f"Vt_k={k}": Vt_k
    }
    excel_path = os.path.join(save_dir, f"{basename}_SVD_matrices.xlsx")
    save_matrix_to_excel(matrices, excel_path)

    return result_path


# === Класс приложения ===

class SVDCompressorApp:
    def __init__(self, master):
        self.master = master
        master.title("SVD Compressor (Images & Excel)")
        master.geometry("700x600")

        self.tabs = ttk.Notebook(master)
        self.main_tab = Frame(self.tabs)
        self.help_tab = Frame(self.tabs)

        self.tabs.add(self.main_tab, text="Главная")
        self.tabs.add(self.help_tab, text="Как пользоваться")
        self.tabs.pack(expand=1, fill="both")

        self.label = Label(self.main_tab, text="Выберите файл (изображение или Excel)", font=("Arial", 14))
        self.label.pack(pady=10)

        self.load_button = Button(self.main_tab, text="Открыть файл", command=self.load_file)
        self.load_button.pack(pady=5)

        self.k_scale = Scale(
            self.main_tab, from_=1, to=100, orient=HORIZONTAL, label="Количество компонент (k)"
        )
        self.k_scale.set(20)
        self.k_scale.pack(pady=10, fill='x', padx=40)

        self.compress_button = Button(self.main_tab, text="Сжать и сохранить", command=self.compress_and_save)
        self.compress_button.pack(pady=5)

        self.preview_label = Label(self.main_tab)
        self.preview_label.pack(pady=10)

        help_text = (
            "📘 Инструкция по использованию:\n\n"
            "1️⃣ Нажмите «Открыть файл» и выберите изображение (.jpg, .png) или Excel (.xlsx).\n"
            "2️⃣ Установите количество компонент SVD (k).\n"
            "3️⃣ Нажмите «Сжать и сохранить».\n\n"
            "📂 Для изображений создаётся папка с:\n"
            "   - исходником, сжатым файлом и Excel-матрицами.\n\n"
            "📊 Для Excel сохраняются:\n"
            "   - восстановленная таблица и матрицы U, S, Vt.\n\n"
            "💡 Чем меньше k, тем сильнее сжатие (но ниже точность).\n"
            "\nАвтор: Эдуард Ефимов | Версия 1.1"
        )
        self.help_label = Label(self.help_tab, text=help_text, justify="left", font=("Arial", 11), wraplength=650)
        self.help_label.pack(padx=10, pady=10)

        self.image = None
        self.file_path = None
        self.is_excel = False

    def load_file(self):
        self.file_path = filedialog.askopenfilename(
            title="Выберите файл",
            filetypes=[
                ("Поддерживаемые файлы", "*.jpg *.jpeg *.png *.bmp *.tif *.tiff *.xlsx *.xls"),
                ("Изображения", "*.jpg *.jpeg *.png *.bmp *.tif *.tiff"),
                ("Excel", "*.xlsx *.xls")
            ]
        )
        if not self.file_path:
            return

        ext = os.path.splitext(self.file_path)[1].lower()
        self.is_excel = ext in [".xlsx", ".xls"]

        if self.is_excel:
            self.label.config(text=f"Загружен Excel файл: {os.path.basename(self.file_path)}")
            self.preview_label.config(image='', text="(Предпросмотр таблицы не поддерживается)")
        else:
            self.image = io.imread(self.file_path)
            img_preview = Image.open(self.file_path)
            img_preview.thumbnail((300, 300))
            img_tk = ImageTk.PhotoImage(img_preview)
            self.preview_label.configure(image=img_tk)
            self.preview_label.image = img_tk
            self.label.config(text=f"Загружено изображение: {os.path.basename(self.file_path)}")

        if not self.is_excel:
            if self.image.ndim == 2:
                m, n = self.image.shape
            else:
                m, n, _ = self.image.shape
        else:
            wb = load_workbook(self.file_path, data_only=True)
            ws = wb.active
            m, n = ws.max_row, ws.max_column

        self.k_scale.config(to=min(m, n))
        self.k_scale.set(min(20, m, n))


    def compress_and_save(self):
        if not self.file_path:
            messagebox.showwarning("Ошибка", "Сначала выберите файл.")
            return

        k = self.k_scale.get()
        basename = os.path.splitext(os.path.basename(self.file_path))[0]
        output_dir = f"compressed_{basename}"
        os.makedirs(output_dir, exist_ok=True)

        if self.is_excel:
            # === Сохраняем оригинальный Excel ===
            original_excel_path = os.path.join(output_dir, f"{basename}_original.xlsx")
            try:
                import shutil
                shutil.copy2(self.file_path, original_excel_path)
            except Exception as e:
                messagebox.showwarning("Ошибка", f"Не удалось скопировать оригинальный Excel:\n{e}")

            # === Компрессия и сохранение результатов ===
            compressed_path = compress_excel_svd(self.file_path, k, output_dir, basename)
            messagebox.showinfo(
                "Готово",
                f"📊 Excel обработан!\n\n"
                f"✅ Оригинальный файл сохранён:\n{original_excel_path}\n"
                f"✅ Сжатая таблица сохранена:\n{compressed_path}\n"
                f"✅ Матрицы U, S, Vt сохранены в Excel-файле\n(в той же папке)"
            )

        else:
            # === Для изображений ===
            original_path = os.path.join(output_dir, f"{basename}_original.jpg")
            io.imsave(original_path, self.image)

            compressed_img = compress_image_svd(self.image, k, output_dir, basename)
            compressed_path = os.path.join(output_dir, f"{basename}_compressed_k={k}.jpg")
            io.imsave(compressed_path, compressed_img)

            messagebox.showinfo(
                "Готово",
                f"🖼️ Изображение обработано!\n\n"
                f"✅ Оригинал сохранён: {original_path}\n"
                f"✅ Сжатое изображение сохранено: {compressed_path}\n"
                f"✅ Матрицы сохранены в Excel-файлы (в той же папке)"
            )



# === Запуск ===
if __name__ == "__main__":
    root = Tk()
    app = SVDCompressorApp(root)
    root.mainloop()
