import numpy as np
from skimage import io
from tkinter import Tk, filedialog, Label, Button, Scale, HORIZONTAL, messagebox, ttk, Frame
from PIL import Image, ImageTk
from openpyxl import Workbook
import os

# === Функции работы с матрицами и Excel ===

def save_matrix_to_excel(matrices, filename):
    """Сохраняет словарь {имя: матрица} в Excel файл"""
    wb = Workbook()
    for name, matrix in matrices.items():
        ws = wb.create_sheet(title=name[:31])  # Excel ограничение — 31 символ
        m, n = matrix.shape
        for i in range(m):
            row_values = matrix[i, :].tolist()
            ws.append(row_values)
    # Удаляем дефолтный лист
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    wb.save(filename)


def compress_image_svd(img, k, save_dir, basename):
    """Компрессия изображения через SVD + сохранение матриц в Excel"""
    if img.ndim == 2:
        num_channels = 1
    else:
        num_channels = img.shape[2]

    new_img = []
    for i in range(3 if num_channels > 1 else 1):
        channel = img if num_channels == 1 else img[:, :, i]
        U, S, Vt = np.linalg.svd(channel, full_matrices=False)

        # Обрезаем до k компонент
        U_k = U[:, :k]
        S_k = np.diag(S[:k])
        Vt_k = Vt[:k, :]

        # === Сохраняем все матрицы в Excel ===
        matrices = {
            f"U_full_ch{i+1}": U,
            f"S_full_ch{i+1}": np.diag(S),
            f"Vt_full_ch{i+1}": Vt,
            f"U_k={k}_ch{i+1}": U_k,
            f"S_k={k}_ch{i+1}": S_k,
            f"Vt_k={k}_ch{i+1}": Vt_k
        }

        excel_path = os.path.join(save_dir, f"{basename}_channel{i+1}_matrices.xlsx")
        save_matrix_to_excel(matrices, excel_path)

        # === Восстанавливаем канал ===
        compressed_channel = U_k @ S_k @ Vt_k
        new_img.append(compressed_channel)

    img_compressed = np.stack(new_img, axis=2) if num_channels > 1 else new_img[0]
    img_compressed = np.clip(img_compressed, 0, 255).astype(np.uint8)
    return img_compressed


# === Класс приложения ===

class SVDCompressorApp:
    def __init__(self, master):
        self.master = master
        master.title("SVD Image Compressor")
        master.geometry("700x600")

        # Вкладки (Notebook)
        self.tabs = ttk.Notebook(master)
        self.main_tab = Frame(self.tabs)
        self.help_tab = Frame(self.tabs)

        self.tabs.add(self.main_tab, text="Главная")
        self.tabs.add(self.help_tab, text="Как пользоваться")
        self.tabs.pack(expand=1, fill="both")

        # --- Главная вкладка ---
        self.label = Label(self.main_tab, text="Выберите изображение для сжатия", font=("Arial", 14))
        self.label.pack(pady=10)

        self.load_button = Button(self.main_tab, text="Открыть изображение", command=self.load_image)
        self.load_button.pack(pady=5)

        self.k_scale = Scale(self.main_tab, from_=1, to=100, orient=HORIZONTAL, label="Количество компонент (k)")
        self.k_scale.set(20)
        self.k_scale.pack(pady=10, fill='x', padx=40)

        self.compress_button = Button(self.main_tab, text="Сжать и сохранить", command=self.compress_and_save)
        self.compress_button.pack(pady=5)

        self.preview_label = Label(self.main_tab)
        self.preview_label.pack(pady=10)

        # --- Вкладка помощи ---
        help_text = (
            "📘 Инструкция по использованию:\n\n"
            "1️⃣ Нажмите «Открыть изображение» и выберите файл (JPG, PNG, BMP и т.д.)\n"
            "2️⃣ Выберите количество компонент SVD (ползунок k). Чем меньше k, тем сильнее сжатие.\n"
            "3️⃣ Нажмите «Сжать и сохранить».\n\n"
            "📂 Программа создаст папку:\n"
            "   compressed_images_<имя_файла>\n\n"
            "В ней будут сохранены:\n"
            " - исходное изображение\n"
            " - сжатое изображение (compressed_k=...jpg)\n"
            " - Excel-файлы с матрицами U, S, Vt для каждого канала RGB\n\n"
            "💡 Чем больше k — тем выше качество и размер файла.\n"
            "💾 Матрицы сохраняются полностью, их можно открывать в Excel для анализа.\n"
            "\nАвтор: Эдуард Ефимов | Версия 1.0"
        )

        self.help_label = Label(self.help_tab, text=help_text, justify="left", font=("Arial", 11), wraplength=650)
        self.help_label.pack(padx=10, pady=10)

        # --- Внутренние переменные ---
        self.image = None
        self.img_path = None

    def load_image(self):
        self.img_path = filedialog.askopenfilename(
            title="Выберите изображение",
            filetypes=[("Изображения", "*.jpg *.jpeg *.png *.bmp *.tif *.tiff")]
        )
        if self.img_path:
            self.image = io.imread(self.img_path)
            img_preview = Image.open(self.img_path)
            img_preview.thumbnail((300, 300))
            img_tk = ImageTk.PhotoImage(img_preview)
            self.preview_label.configure(image=img_tk)
            self.preview_label.image = img_tk
            self.label.config(text=f"Загружено: {os.path.basename(self.img_path)}")

    def compress_and_save(self):
        if self.image is None:
            messagebox.showwarning("Ошибка", "Сначала выберите изображение.")
            return

        k = self.k_scale.get()
        basename = os.path.splitext(os.path.basename(self.img_path))[0]

        # Папка для сохранения
        output_dir = f"compressed_images_{basename}"
        os.makedirs(output_dir, exist_ok=True)

        # Сохраняем оригинальное изображение
        original_save_path = os.path.join(output_dir, f"{basename}_original.jpg")
        io.imsave(original_save_path, self.image)

        # Компрессия и сохранение Excel-файлов
        compressed_img = compress_image_svd(self.image, k, output_dir, basename)

        # Сохраняем сжатое изображение
        compressed_save_path = os.path.join(output_dir, f"{basename}_compressed_k={k}.jpg")
        io.imsave(compressed_save_path, compressed_img)

        messagebox.showinfo(
            "Готово",
            f"✅ Оригинальное изображение сохранено:\n{original_save_path}\n"
            f"✅ Сжатое изображение сохранено:\n{compressed_save_path}\n"
            f"✅ Матрицы сохранены в Excel-файлы в папке:\n{output_dir}"
        )


# === Запуск ===
if __name__ == "__main__":
    root = Tk()
    app = SVDCompressorApp(root)
    root.mainloop()
